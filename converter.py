import pandas as pd
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from telemetry import flatten_record


def flatten_json(data, parent_key="", sep="."):
    """
    Generic JSON flattener for non-telemetry JSON.
    """

    items = {}

    if isinstance(data, dict):

        for key, value in data.items():

            new_key = f"{parent_key}{sep}{key}" if parent_key else key

            if isinstance(value, dict):

                items.update(flatten_json(value, new_key, sep))

            elif isinstance(value, list):

                if len(value) == 0:
                    items[new_key] = ""

                else:
                    for i, item in enumerate(value):
                        items.update(
                            flatten_json(
                                item,
                                f"{new_key}[{i}]",
                                sep
                            )
                        )

            else:

                items[new_key] = value

    return items


def export_to_excel(json_object):

    try:

        # -------------------------
        # Build dataframe
        # -------------------------

        if isinstance(json_object, dict) and "items" in json_object:

            rows = []

            for record in json_object["items"]:
                rows.append(flatten_record(record))

            df = pd.DataFrame(rows)

        else:

            df = pd.DataFrame([flatten_json(json_object)])

        # -------------------------
        # Ask save location
        # -------------------------

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            title="Save Excel File"
        )

        if not filename:
            return

        # -------------------------
        # Export
        # -------------------------

        with pd.ExcelWriter(
            filename,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="Telemetry",
                index=False
            )

            ws = writer.sheets["Telemetry"]

            # -------------------------
            # Header formatting
            # -------------------------

            header_fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            header_font = Font(
                bold=True,
                color="FFFFFF"
            )

            for cell in ws[1]:

                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # -------------------------
            # Freeze first row
            # -------------------------

            ws.freeze_panes = "A2"

            # -------------------------
            # Auto Filter
            # -------------------------

            ws.auto_filter.ref = ws.dimensions

            # -------------------------
            # Auto Column Width
            # -------------------------

            for column in ws.columns:

                length = max(
                    len(str(cell.value)) if cell.value else 0
                    for cell in column
                )

                ws.column_dimensions[
                    get_column_letter(column[0].column)
                ].width = min(length + 3, 50)

        messagebox.showinfo(
            "Success",
            f"Excel exported successfully!\n\n{filename}"
        )

    except Exception as e:

        messagebox.showerror(
            "Export Error",
            str(e)
        )